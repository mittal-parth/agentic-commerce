"""
Thin FastAPI wrapper around onboard_merchant.py.
Saves uploaded catalogue to a temp file and invokes the CLI via subprocess.
"""

import csv
import io
import re
import subprocess
import sys
import tempfile
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

# Repo root (parent of web/backend)
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
ONBOARD_SCRIPT = REPO_ROOT / "onboard_merchant.py"
DEPLOY_DIR = REPO_ROOT / "deploy"

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

app = FastAPI(
    title="Merchant Onboarding API",
    description="Upload a catalogue and merchant details to generate a UCP-ready data package.",
)

# CORS so Vite dev server can call the API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def slugify(name: str) -> str:
    """Derive a filesystem-safe slug from merchant name (e.g. 'Artisan India' -> 'artisan-india')."""
    slug = re.sub(r"[^\w\s-]", "", name.lower())
    slug = re.sub(r"[-\s]+", "-", slug).strip("-")
    return slug or "merchant"


def xlsx_to_csv_bytes(content: bytes) -> bytes:
    """Convert first sheet of an Excel workbook to CSV bytes."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)
    wb.close()
    return buf.getvalue().encode("utf-8")


@app.post("/api/onboard")
async def onboard(
    merchant_name: str = Form(..., description="Merchant display name"),
    merchant_vpa: str = Form(..., description="Merchant UPI VPA (e.g. shop@ybl)"),
    catalogue: UploadFile = File(..., description="Catalogue CSV or Excel file"),
) -> dict:
    """Run onboard_merchant.py with the uploaded file and form fields."""
    suffix = Path(catalogue.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid file type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    content = await catalogue.read()
    if not content:
        raise HTTPException(status_code=422, detail="Catalogue file is empty.")

    if suffix == ".xlsx":
        content = xlsx_to_csv_bytes(content)

    with tempfile.NamedTemporaryFile(
        mode="wb", suffix=".csv", delete=False
    ) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        output_slug = slugify(merchant_name)
        output_dir = DEPLOY_DIR / output_slug
        output_dir.mkdir(parents=True, exist_ok=True)

        result = subprocess.run(
            [
                sys.executable,
                str(ONBOARD_SCRIPT),
                "--catalogue",
                tmp_path,
                "--merchant-name",
                merchant_name,
                "--merchant-vpa",
                merchant_vpa,
                "--output-dir",
                str(output_dir),
            ],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if result.returncode != 0:
        raise HTTPException(
            status_code=422,
            detail=result.stderr.strip() or result.stdout.strip() or "Onboarding failed.",
        )

    return {
        "status": "ok",
        "merchant_name": merchant_name,
        "output_dir": str(output_dir),
    }
